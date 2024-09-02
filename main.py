from PyQt5.QtWidgets import QMainWindow, QApplication, QFileDialog, QWidget, QTabWidget, QPlainTextEdit, QSizePolicy
from PyQt5.QtSerialPort import QSerialPort, QSerialPortInfo
from PyQt5.QtCore import QIODevice
from PyQt5 import uic
import sys
import datetime
import openpyxl
from channel import Channel
from general_view import GeneralView


# Creating the main window
class Main(QMainWindow):
    def __init__(self):
        super().__init__()
        uic.loadUi("main.ui", self)
        self.title = "Коагулограф"
        self.left = 500
        self.top = 300
        self.setWindowTitle(self.title)
        self.maxTopValue = 400
        self.maxRightValue = 2000
        self.chChars = ['@', '#', '$', '%', '^']

        # Initialize tab screen

        self.strok_data = ''
        self.oldstrok_data = ''
        self.output_rate = 10
        self.counter_rate = 0
        self.setWindowTitle("Коагулограф")
        self.dt_now = datetime.datetime.today()
        self.dateTimeEdit.setDateTime(self.dt_now)
        self.interferences = 0
        self.connectButton.clicked.connect(self.buttonConDis)
        self.saveButton.clicked.connect(self.save)
        self.clearButton.clicked.connect(self.onClear)
        self.importButton.clicked.connect(self.onImport)
        self.serial = QSerialPort()
        self.serial.setBaudRate(9600)
        self.isConnected = False
        self.wasStopped = True
        self.ports_name_list = []
        self.ports_num_list = []
        ports = QSerialPortInfo().availablePorts()
        for port in ports:
            full_name = port.portName() + " " + port.description()
            self.ports_name_list.append(full_name)
            self.ports_num_list.append(port.portName())

        self.ports.addItems(self.ports_name_list)
        self.serial.readyRead.connect(self.onRead)
        self.graph_data = []
        self.par_avg = []
        self.name_data_patient = ["Дата и время", "Доп. время(сек)", "ФИО", "№ Истории болезни",
                                  "Диагноз", "Препараты", "Обстоятельства", "Фибриноген", "ПТИ", "МНО", "АЧТВ", "ACT",
                                  "Д-Димер",
                                  "Тромбоциты"]
        self.named_graph_data = ["Время до начала свёртывания, сек",
                                 "Время до окончания свёртывания, сек",
                                 "Длительность свёртывания, сек",
                                 "Время до начала ретракции, сек",
                                 "Время до окончания ретракции, сек",
                                 "Длительность ретракции, сек",
                                 "Максимальная амплитуда, ед",
                                 "Минимальная амплитуда, ед",
                                 "Амплитуда на 3 минуте фибринолиза, ед",
                                 "Скорость свёртывания, ед/мин",
                                 "Скорость нарастания фибринолиза, ед/мин",
                                 "Коэффицент ректракции, %",
                                 "Коэффицент фибринолиза, %",
                                 "Активность фибринолиза, %"]

        self.tabs = QTabWidget()
        self.tab1 = Channel(self)
        self.tab2 = Channel(self)
        self.tab3 = Channel(self)
        self.tab4 = Channel(self)
        self.tab5 = Channel(self)
        self.tab6 = GeneralView(self)
        self.tab7 = QPlainTextEdit()
        self.tabsList = [self.tab1, self.tab2, self.tab3, self.tab4, self.tab5]

        font = self.tab1.output.font()
        font.setPointSize(14)
        self.tab7.setFont(font)
        # Add tabs
        self.tabs.addTab(self.tab1, "Channel 1")
        self.tabs.addTab(self.tab2, "Channel 2")
        self.tabs.addTab(self.tab3, "Channel 3")
        self.tabs.addTab(self.tab4, "Channel 4")
        self.tabs.addTab(self.tab5, "Channel 5")
        self.tabs.addTab(self.tab6, "General view")
        self.tabs.addTab(self.tab7, "General info")

        self.tabsLayout.addWidget(self.tabs)
        self.tabs.sizePolicy().setHorizontalPolicy(QSizePolicy.Expanding)
        self.centralWidget = QWidget()
        self.centralWidget.setLayout(self.mainLayout)
        self.setCentralWidget(self.centralWidget)
        self.show()

    def buttonConDis(self):
        if self.connectButton.text() == "Начать":
            self.connectButton.setText("Остановить")
            self.onConnect()
        elif self.connectButton.text() == "Остановить":
            self.wasStopped = True
            self.connectButton.setText("Начать")
            self.onDisconnect()

    def updateGlobalInfo(self):
        params_avg = [0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0]
        for tab in self.tabsList:
            if self.tab6.check_boxes[self.tabsList.index(tab)]:
                for par_i in range(len(tab.tab_param)):
                    params_avg[par_i] += tab.tab_param[par_i]
        for i in range(len(params_avg)):
            params_avg[i] = round(params_avg[i] / sum(self.tab6.check_boxes), 2)
        text = ""
        text = "Время до начала свёртывания, сек" + " " * (40 - len("Время до начала свёртывания, сек")) + str(params_avg[0]) + " " * (10 - len(str(params_avg[0]))) + self.tab1.secToMin(params_avg[0]) + "\n"
        text += "Время до окончания свёртывания, сек" + " " * (40 - len("Время до окончания свёртывания, сек")) + str(params_avg[1]) + " " * (10 - len(str(params_avg[1]))) + self.tab1.secToMin(params_avg[1]) + "\n"
        text += "Длительность свёртывания, сек" + " " * (40 - len("Длительность свёртывания, сек")) + str(params_avg[2]) + " " * (10 - len(str(params_avg[2]))) + self.tab1.secToMin(params_avg[2]) + "\n"
        text += "Время до начала ретракции, сек" + " " * (40 - len("Время до начала ретракции, сек")) + str(params_avg[3]) + " " * (10 - len(str(params_avg[3]))) + self.tab1.secToMin(params_avg[3]) + "\n"
        text += "Время до окончания ретракции, сек" + " " * (40 - len("Время до окончания ретракции, сек")) + str(params_avg[4]) + " " * (10 - len(str(params_avg[4]))) + self.tab1.secToMin(params_avg[4]) + "\n"
        text += "Длительность ретракции, сек" + " " * (40 - len("Длительность ретракции, сек")) + str(params_avg[5]) + " " * (10 - len(str(params_avg[5]))) + self.tab1.secToMin(params_avg[5]) + "\n"

        text += "Максимальная амплитуда, ед" + " " * (40 - len("Максимальная амплитуда, ед")) + str(params_avg[6]) + "\n"
        text += "Минимальная амплитуда, ед" + " " * (40 - len("Минимальная амплитуда, ед")) + str(params_avg[7]) + "\n"
        text += "Амплитуда на 3 минуте фибринолиза, ед" + " " * (40 - len("Амплитуда на 3 минуте фибринолиза, ед")) + str(params_avg[8]) + "\n"

        text += "Скорость свёртывания, ед/мин" + " " * (40 - len("Скорость свёртывания, ед/мин")) + str(params_avg[9]) + "\n"
        text += "Скорость нарастания фибринолиза, ед/мин" + " " * (40 - len("Скорость нарастания фибринолиза, ед/мин")) + str(params_avg[10]) + "\n"

        text += "Коэффициент ректракции, %" + " " * (40 - len("Коэффициент ректракции, %")) + str(params_avg[11]) + "\n"
        text += "Коэффициент фибринолиза, %" + " " * (40 - len("Коэффициент фибринолиза, %")) + str(params_avg[12]) + "\n"
        text += "Активность фибринолиза, %" + " " * (40 - len("Активность фибринолиза, %")) + str(params_avg[13]) + "\n"
        self.par_avg = params_avg
        self.tab7.setPlainText(text)


    def onConnect(self):
        print("connect")
        try:
            choose_index = self.ports_name_list.index(self.ports.currentText())
            choose_com_port = self.ports_num_list[choose_index]
            self.serial.setPortName(choose_com_port)
            self.serial.open(QIODevice.ReadWrite)
            self.serial.readyRead.connect(self.onRead)
            self.sendData('a' + str(self.alpha1.value()))
            self.sendData('b' + str(self.alpha2.value()))
            self.sendData('r')
        except Exception as err:
            print(err)

    def onDisconnect(self):
        self.sendData('p')
        # self.serial.close()

    def onClear(self):
        for tab in self.tabsList:
            tab.chClear()
        self.tab6.clear()
        self.onDisconnect()
        self.interferences = 0
        self.oldstrok_data = ''
        self.strok_data = ''
        self.data_list = []
        self.norm_data_list = []
        self.time = []
        self.graph_data = []
        self.nameEdit.setText('')
        self.numEdit.setText('')
        self.diagnosisEdit.clear()
        self.conditionEdit.clear()
        self.dt_now = datetime.datetime.today()
        self.dateTimeEdit.setDateTime(self.dt_now)
        self.fibrinogenEdit.setValue(0)
        self.ptiEdit.setValue(0)
        self.mnoEdit.setValue(0)
        self.actvEdit.setValue(0)
        self.actEdit.setValue(0)
        self.addTimeEdit.setValue(0)
        self.ddimerEdit.setValue(0)
        self.trombEdit.setValue(0)
        self.medicationEdit.clear()

    def onRead(self):
        try:
            while self.serial.bytesAvailable() > 0:
                data = self.serial.readLine()
                self.strok_data = str(data)[2:-1]
                if r"\n" in self.strok_data:
                    self.oldstrok_data += self.strok_data
                    num = self.oldstrok_data[1:-4]
                    chChar = self.oldstrok_data[0]
                    self.oldstrok_data = ''
                    if chChar in self.chChars:
                        if int(num) > self.maxTopValue:
                            num = self.maxTopValue
                        self.tabsList[self.chChars.index(chChar)].writeData(round(int(num) / self.maxTopValue * 100, 2))
                        self.tab6.update_data(self.chChars.index(chChar), round(int(num) / self.maxTopValue * 100, 2))
        except Exception as err:
            print("err", err)

    def sendData(self, data_to_send):
        if self.serial.isOpen():
            data_to_send += "*"
            self.serial.write(data_to_send.encode())

    def save(self):
        data_patient = [self.dateTimeEdit.dateTime().toString('dd.MM.yyyy hh:mm'), self.addTimeEdit.value(),
                        self.nameEdit.text(), self.numEdit.text(), self.diagnosisEdit.toPlainText(),
                        self.conditionEdit.toPlainText(), self.medicationEdit.toPlainText(),
                        self.fibrinogenEdit.value(), self.ptiEdit.value(), self.mnoEdit.value(), self.actvEdit.value(),
                        self.actEdit.value(), self.ddimerEdit.value(), self.trombEdit.value()]

        wb = openpyxl.Workbook()
        wb.create_sheet(title='Лист', index=0)
        counter = 1
        for tab in self.tabsList:
            sheet = wb['Лист']
            for row in range(len(tab.chanelTime)):
                cell = sheet.cell(row=row + 1, column=counter)
                cell.value = tab.chanelTime[row]
            counter += 1
            for row in range(len(tab.chanelData)):
                cell = sheet.cell(row=row + 1, column=counter)
                cell.value = tab.chanelData[row]
            counter += 1

        all_name_data = self.name_data_patient + self.named_graph_data
        for row in range(len(all_name_data)):
            cell = sheet.cell(row=row + 1, column=11)
            cell.value = all_name_data[row]

        all_data = data_patient + self.graph_data
        for row in range(len(all_data)):
            cell = sheet.cell(row=row + 1, column=12)
            cell.value = all_data[row]

        for row in range(len(self.named_graph_data)):
            cell = sheet.cell(row=row + 2, column=13)
            cell.value = self.named_graph_data[row]

        cell = sheet.cell(row=1, column=14)
        cell.value = "Channel 1"
        cell = sheet.cell(row=1, column=15)
        cell.value = "Channel 2"
        cell = sheet.cell(row=1, column=16)
        cell.value = "Channel 3"
        cell = sheet.cell(row=1, column=17)
        cell.value = "Channel 4"
        cell = sheet.cell(row=1, column=18)
        cell.value = "Channel 5"
        cell = sheet.cell(row=1, column=19)
        cell.value = "Average"

        for row in range(len(self.tabsList[0].tab_param)):
            cell = sheet.cell(row=row + 2, column=14)
            cell.value = self.tabsList[0].tab_param[row]

        for row in range(len(self.tabsList[1].tab_param)):
            cell = sheet.cell(row=row + 2, column=15)
            cell.value = self.tabsList[1].tab_param[row]

        for row in range(len(self.tabsList[2].tab_param)):
            cell = sheet.cell(row=row + 2, column=16)
            cell.value = self.tabsList[2].tab_param[row]

        for row in range(len(self.tabsList[3].tab_param)):
            cell = sheet.cell(row=row + 2, column=17)
            cell.value = self.tabsList[3].tab_param[row]

        for row in range(len(self.tabsList[4].tab_param)):
            cell = sheet.cell(row=row + 2, column=18)
            cell.value = self.tabsList[4].tab_param[row]

        for row in range(len(self.par_avg)):
            cell = sheet.cell(row=row + 2, column=19)
            cell.value = self.par_avg[row]

        try:
            filename = QFileDialog.getSaveFileName(self, "Сохранить в таблицу",
                                                   str(self.nameEdit.text().split()[0]) + '_' +
                                                   self.dateTimeEdit.dateTime().toString('dd-MM-yyyy_hh-mm'), "*.xlsx")
        except:
            filename = QFileDialog.getSaveFileName(self, "Сохранить в таблицу", '', "*.xlsx")
        try:
            wb.save(filename[0])
        except:
            print('Save error')

    def onImport(self):
        self.onClear()
        filename = QFileDialog.getOpenFileName(self, "Импортировать из таблицы", '', "*.xlsx")

        wb = openpyxl.load_workbook(filename[0])
        sh_names = wb.sheetnames
        # print(sh_names, filename)
        sheet = wb[sh_names[0]]
        column = 1
        try:
            for tab in self.tabsList:
                row = 1
                while True:
                    row += 1
                    cell = sheet.cell(row=row, column=column)
                    val = cell.value
                    if val == None:
                        break
                    else:
                        tab.chanelTime.append(val)
                        cell = sheet.cell(row=row, column=column + 1)
                        val = cell.value  # / self.maxTopValue * 100
                        # print(val)
                        tab.chanelData.append(val)
                column += 2
                tab.chImport()
                self.tab6.generalImport(self.tabsList.index(tab), tab.chanelTime, tab.chanelData)
        except Exception as err:
            print(err)
        # norm_data_list = [int(item / self.maxTopValue * 100) for item in self.data_list]
        # self.graph.plot(self.time, norm_data_list, pen=self.pen)
        # self.graph.disableAutoRange()
        # self.graph.setLimits(yMin=-10, yMax=100, xMin=0, xMax=self.maxRightValue)
        try:
            data_patient = []
            for row in range(len(self.name_data_patient)):
                cell = sheet.cell(row=row + 1, column=12)
                val = cell.value
                data_patient.append(val)
            datetime_str1 = data_patient[0]
            self.dateTimeEdit.setDateTime(datetime.datetime.strptime(datetime_str1, '%d.%m.%Y %H:%M'))
            self.addTimeEdit.setValue(data_patient[1])
            self.nameEdit.setText(data_patient[2])
            self.numEdit.setText(data_patient[3])
            self.diagnosisEdit.setPlainText(data_patient[4])
            self.conditionEdit.setPlainText(data_patient[5])
            self.medicationEdit.setPlainText(data_patient[6])
            self.fibrinogenEdit.setValue(data_patient[7])
            self.ptiEdit.setValue(data_patient[8])
            self.mnoEdit.setValue(data_patient[9])
            self.actvEdit.setValue(data_patient[10])
            self.actEdit.setValue(data_patient[11])
            self.ddimerEdit.setValue(data_patient[12])
            self.trombEdit.setValue(data_patient[13])
        except Exception as err:
            print(err)

    def calculate(self):
        self.tab1.chCalculate(self)
        self.tab2.chCalculate(self)
        self.tab3.chCalculate(self)
        self.tab4.chCalculate(self)
        self.tab5.chCalculate(self)
        self.updateGlobalInfo()


if __name__ == '__main__':
    app = QApplication(sys.argv)
    ex = Main()
    ex.serial.close()
    sys.exit(app.exec_())
